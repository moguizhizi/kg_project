# utils/xlsx_utils.py

import json
import os
import pandas as pd
import zipfile
from xml.etree import ElementTree
from typing import List
from pathlib import Path
from typing import Dict

BASE_DIR = Path(__file__).resolve().parent
ONTOLOGY_MAPPINGS_DIR = BASE_DIR / "ontology_mappings"
HOME_BASED_USER_TRAINING = BASE_DIR / "entity_registry" / "home_based_user_training"


def build_column_mapping(
    excel_to_label: Dict[str, str],
    property_ontology: Dict[str, str],
    entity_ontology: Dict[str, str],
    strict: bool = True,
) -> Dict[str, str]:
    """
    Build XLSX column mapping WITHOUT reversing ontology dicts.

    Excel column -> AU_P / AU_Q
    """

    column_mapping: Dict[str, str] = {}

    for excel_col, label in excel_to_label.items():
        label = str(label).strip()
        resolved_label = None

        # 1. 在属性表中顺序查
        for au_code, au_label in property_ontology.items():
            if au_code == label:
                resolved_label = au_label
                break

        # 2. 如果属性没找到，再查实体表
        if resolved_label is None:
            for au_code, au_label in entity_ontology.items():
                if au_code == label:
                    resolved_label = au_label
                    break

        # 3. 处理结果
        if resolved_label:
            column_mapping[excel_col] = resolved_label
        else:
            if strict:
                raise ValueError(
                    f"Cannot resolve column '{excel_col}' "
                    f"(label='{label}') to AU_P or AU_Q"
                )

    return column_mapping


from openpyxl import load_workbook
from typing import List


def load_unique_column_fast(xlsx_path: str, sheet_name: str, column_name: str) -> List:
    """
    超大 XLSX 文件快速读取指定列，去空值、去重，返回列表。

    参数:
        xlsx_path: Excel 文件路径
        sheet_name: 需要读取的 sheet 名
        column_name: 指定列名

    返回:
        List: 去空、去重后的列值列表
    """
    # 打开 workbook，read_only 模式
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb[sheet_name]

    # 找到列索引（从 0 开始）
    col_idx = None
    for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1))):
        if cell.value == column_name:
            col_idx = i
            break
    if col_idx is None:
        raise ValueError(f"列 {column_name} 不存在")

    # 流式读取列
    seen = set()
    unique_values = []
    for row in ws.iter_rows(min_row=2):
        val = row[col_idx].value
        if val is not None and val not in seen:
            seen.add(val)
            unique_values.append(val)

    wb.close()
    return unique_values


def load_unique_column(
    path: str, sheet_name: str, column_name: str, as_list: bool = False
) -> pd.DataFrame | List:
    """
    读取 XLSX 文件中指定列，去空、去重，并返回整理后的结果。

    参数:
        path: Excel 文件路径
        sheet_name: 需要读取的 sheet 名
        column_name: 指定列名
        as_list: 是否返回列表 (默认 False 返回 DataFrame)

    返回:
        去空、去重后的 DataFrame 或列表
    """
    # 只读取指定列
    df = pd.read_excel(
        path, sheet_name=sheet_name, usecols=[column_name], engine="openpyxl"
    )

    # 去掉空值
    df = df.dropna(subset=[column_name])

    # 去重
    df = df.drop_duplicates(subset=[column_name])

    # 重置索引
    df = df.reset_index(drop=True)

    if as_list:
        return df[column_name].tolist()
    return df


def get_xlsx_sheetnames(xlsx_path: str) -> List[str]:
    """
    快速获取 XLSX 文件的所有 sheet 名，不读取数据

    参数:
        xlsx_path: Excel 文件路径

    返回:
        List[str]: sheet 名列表
    """
    with zipfile.ZipFile(xlsx_path) as z:
        wb_xml = z.read("xl/workbook.xml")
        root = ElementTree.fromstring(wb_xml)
        ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheet_names = [s.attrib["name"] for s in root.findall(".//ns:sheet", ns)]
    return sheet_names


# =========================
# main (demo / test)
# =========================
def main():

    with open(os.path.join(ONTOLOGY_MAPPINGS_DIR, "entity_type2label.json"), "r") as f:
        ENTITY_TYPE_2_LABEL = json.load(f)

    with open(os.path.join(ONTOLOGY_MAPPINGS_DIR, "prop2label.json"), "r") as f:
        PROP_2_LABEL = json.load(f)

    with open(os.path.join(HOME_BASED_USER_TRAINING, "column_mapping.json"), "r") as f:
        COLUMN_MAPPING = json.load(f)

    column_mapping = build_column_mapping(
        excel_to_label=COLUMN_MAPPING,
        property_ontology=ENTITY_TYPE_2_LABEL,
        entity_ontology=PROP_2_LABEL,
        strict=True,
    )

    print("Final column_mapping:")
    for k, v in column_mapping.items():
        print(f"  {k} -> {v}")


if __name__ == "__main__":
    main()
